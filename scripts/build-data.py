#!/usr/bin/env python3
"""
Mengubah file sumber "CEKLIS LAPORAN BULANAN.xlsx" menjadi satu dataset JSON
terstruktur yang dikonsumsi oleh dashboard web app (src/data/dashboard-data.json).

Jalankan ulang setiap kali ada export baru dari spreadsheet:
    python3 scripts/build-data.py <path-ke-xlsx> [as-of-date YYYY-MM-DD]
"""
import json
import re
import sys
from datetime import date, datetime

import openpyxl

MONTHS_2026 = [
    ("2026-02", "FEBRUARI 2026", "Februari 2026", 2, 28),
    ("2026-03", "MARET 2026", "Maret 2026", 3, 31),
    ("2026-04", "APRIL 2026", "April 2026", 4, 30),
    ("2026-05", "MIE 2026", "Mei 2026", 5, 31),
    ("2026-06", "JUNI 2026", "Juni 2026", 6, 30),
    ("2026-07", "JULI 2026", "Juli 2026", 7, 31),
    ("2026-08", "AGUSTUS 2026", "Agustus 2026", 8, 31),
]

SUFFIXES = [" SATPAM", " CS & SAT", " CS"]


def base_key(name):
    if not name:
        return None
    n = str(name).strip().upper()
    n = re.sub(r"\s+", " ", n)
    for suf in SUFFIXES:
        if n.endswith(suf):
            n = n[: -len(suf)].strip()
            break
    return n


def norm(name):
    if not name:
        return None
    return re.sub(r"\s+", " ", str(name).strip().upper())


def iso(d):
    if isinstance(d, datetime):
        return d.date().isoformat()
    return None


def load(path):
    return openpyxl.load_workbook(path, data_only=True)


def parse_data_sheet(wb):
    """Sheet DATA: NO | NAMA | DATA(metode) | KETERANGAN -> metode per base_key."""
    ws = wb["DATA"]
    out = {}
    for r in range(5, ws.max_row + 1):
        nama = ws.cell(row=r, column=3).value
        metode = ws.cell(row=r, column=4).value
        if not nama:
            continue
        out[base_key(nama)] = metode
    return out


def parse_excluded(wb):
    ws = wb["TIDAK MENGGUNAKAN ABSENSI & LAP"]
    out = []
    for r in range(5, ws.max_row + 1):
        nama = ws.cell(row=r, column=3).value
        ket = ws.cell(row=r, column=4).value
        if not nama:
            continue
        out.append({"nama": str(nama).strip(), "keterangan": ket})
    return out


def parse_client_master(wb):
    """ALL CLAEN 2026: tipe client (LAMA/BARU) per base_key."""
    ws = wb["ALL CLAEN 2026"]
    tipe = {}
    for r in range(4, ws.max_row + 1):
        nama = ws.cell(row=r, column=3).value
        t = ws.cell(row=r, column=4).value
        if not nama or t not in ("CLIENT LAMA", "CLIENT BARU"):
            continue
        tipe[base_key(nama)] = "LAMA" if t == "CLIENT LAMA" else "BARU"
    return tipe


def parse_month_sheet(wb, sheet_name, has_pic):
    ws = wb[sheet_name]
    rows = []
    for r in range(3, ws.max_row + 1):
        nama = ws.cell(row=r, column=2).value
        if not nama or not isinstance(nama, str) or not nama.strip():
            continue
        if has_pic:
            pic = ws.cell(row=r, column=3).value
            wa = ws.cell(row=r, column=4).value
            lap = ws.cell(row=r, column=5).value
            lap_tgl = ws.cell(row=r, column=6).value
            lap_ket = ws.cell(row=r, column=7).value
            abs_ = ws.cell(row=r, column=8).value
            abs_tgl = ws.cell(row=r, column=9).value
            abs_ket = ws.cell(row=r, column=10).value
            catatan = ws.cell(row=r, column=11).value
        else:
            pic = None
            wa = None
            lap = ws.cell(row=r, column=3).value
            lap_tgl = ws.cell(row=r, column=4).value
            lap_ket = ws.cell(row=r, column=5).value
            abs_ = ws.cell(row=r, column=6).value
            abs_tgl = ws.cell(row=r, column=7).value
            abs_ket = ws.cell(row=r, column=8).value
            catatan = None
        if str(nama).strip().upper() in ("KETERANGAN", "NO"):
            continue
        rows.append(
            {
                "nama": str(nama).strip(),
                "pic": (str(pic).strip() if pic and str(pic).strip() else None),
                "namaGroupWA": (str(wa).strip() if wa and str(wa).strip() else None),
                "laporan": {
                    "selesai": bool(lap) if isinstance(lap, bool) else False,
                    "tanggalSelesai": iso(lap_tgl),
                    "keterangan": (str(lap_ket).strip() if lap_ket and str(lap_ket).strip() else None),
                },
                "absensi": {
                    "selesai": bool(abs_) if isinstance(abs_, bool) else False,
                    "tanggalSelesai": iso(abs_tgl),
                    "keterangan": (str(abs_ket).strip() if abs_ket and str(abs_ket).strip() else None),
                },
                "catatan": (str(catatan).strip() if catatan and str(catatan).strip() else None),
            }
        )
    return rows


def month_status(rec, is_closed):
    lap_ok = rec["laporan"]["selesai"]
    abs_ok = rec["absensi"]["selesai"]
    if lap_ok and abs_ok:
        return "selesai"
    if is_closed:
        return "overdue"
    return "proses"


def main():
    src = sys.argv[1] if len(sys.argv) > 1 else (
        "/root/.claude/uploads/63fd1ee0-57ad-5009-b5f0-1b48cb3ed2bf/721dc0d1-CEKLIS_LAPORAN_BULANAN.xlsx"
    )
    as_of = sys.argv[2] if len(sys.argv) > 2 else "2026-08-22"
    as_of_date = date.fromisoformat(as_of)

    wb = load(src)
    metode_by_key = parse_data_sheet(wb)
    excluded = parse_excluded(wb)
    excluded_keys = {base_key(e["nama"]) for e in excluded}
    tipe_by_key = parse_client_master(wb)

    months_out = []
    monthly_rows = {}
    for code, sheet, label, mnum, mdays in MONTHS_2026:
        has_pic = sheet in ("JULI 2026", "AGUSTUS 2026")
        rows = parse_month_sheet(wb, sheet, has_pic)
        monthly_rows[code] = rows

        month_end = date(2026, mnum, mdays)
        is_closed = as_of_date > month_end

        total = len(rows)
        lap_done = sum(1 for r in rows if r["laporan"]["selesai"])
        abs_done = sum(1 for r in rows if r["absensi"]["selesai"])
        both_done = sum(1 for r in rows if r["laporan"]["selesai"] and r["absensi"]["selesai"])
        overdue_count = sum(
            1
            for r in rows
            if is_closed and not (r["laporan"]["selesai"] and r["absensi"]["selesai"])
        )

        months_out.append(
            {
                "code": code,
                "label": label,
                "isClosed": is_closed,
                "totalLokasi": total,
                "laporanSelesai": lap_done,
                "laporanPct": round(lap_done / total, 4) if total else None,
                "absensiSelesai": abs_done,
                "absensiPct": round(abs_done / total, 4) if total else None,
                "keduanyaSelesai": both_done,
                "overdueCount": overdue_count,
            }
        )

    # --- current roster (latest month with data = Agustus 2026) enriched with
    # tipe client, metode pelaporan, and PIC (fallback to Juli 2026 PIC if kosong)
    current_code = months_out[-1]["code"]
    current_rows = monthly_rows[current_code]
    prior_code = months_out[-2]["code"]
    prior_pic_by_key = {norm(r["nama"]): r["pic"] for r in monthly_rows[prior_code] if r.get("pic")}

    current_month_end = date(2026, MONTHS_2026[-1][3], MONTHS_2026[-1][4])
    is_current_closed = as_of_date > current_month_end

    locations = []
    for r in current_rows:
        key = base_key(r["nama"])
        pic = r["pic"] or prior_pic_by_key.get(norm(r["nama"]))
        status = month_status(r, is_current_closed)
        days_late = None
        if status == "overdue":
            days_late = (as_of_date - current_month_end).days
        locations.append(
            {
                "nama": r["nama"],
                "baseKey": key,
                "tipeClient": tipe_by_key.get(key),
                "metode": metode_by_key.get(key),
                "pic": pic,
                "laporan": r["laporan"],
                "absensi": r["absensi"],
                "catatan": r["catatan"],
                "status": status,
                "hariTelat": days_late,
            }
        )

    # --- heatmap: current roster x semua bulan, join via baseKey (bulan Feb-Jun
    # satu baris mencakup Satpam+CS sekaligus, jadi hasil join bisa "tidak ada data"
    # untuk sub-tim yang baru dipisah mulai Juli)
    heatmap_sites = []
    seen_base = set()
    for loc in locations:
        if loc["baseKey"] in seen_base:
            continue
        seen_base.add(loc["baseKey"])

    for base in sorted(seen_base):
        cells = []
        for m in months_out:
            code = m["code"]
            rows = monthly_rows[code]
            if code in ("2026-07", "2026-08"):
                matches = [r for r in rows if base_key(r["nama"]) == base]
            else:
                matches = [r for r in rows if norm(r["nama"]) == base]
            if not matches:
                cells.append({"month": code, "status": "no-data"})
                continue
            lap_ok = all(x["laporan"]["selesai"] for x in matches)
            abs_ok = all(x["absensi"]["selesai"] for x in matches)
            closed = m["isClosed"]
            if lap_ok and abs_ok:
                st = "selesai"
            elif closed:
                st = "overdue"
            else:
                st = "proses"
            cells.append({"month": code, "status": st})
        heatmap_sites.append({"baseKey": base, "cells": cells})

    # urutkan heatmap: yang paling banyak overdue/proses dulu (paling berisiko)
    risk_weight = {"overdue": 3, "proses": 1, "selesai": 0, "no-data": 0}
    heatmap_sites.sort(key=lambda s: sum(risk_weight[c["status"]] for c in s["cells"]), reverse=True)

    # --- leaderboard PIC: basis ranking pakai bulan CLOSED terakhir yang punya
    # kolom PIC (Juli 2026) supaya statusnya final (selesai/overdue), bukan bulan
    # berjalan (Agustus) yang akan membuat semua orang tampak 0% karena belum
    # jatuh tempo. Beban kerja bulan berjalan tetap ditampilkan terpisah.
    juli_rows = monthly_rows.get("2026-07", [])
    juli_month_end = date(2026, 7, 31)
    pic_stats = {}
    for r in juli_rows:
        if not r["pic"]:
            continue
        st = month_status(r, is_closed=True)
        p = pic_stats.setdefault(
            r["pic"], {"pic": r["pic"], "lokasiDitangani": 0, "selesai": 0, "overdue": 0}
        )
        p["lokasiDitangani"] += 1
        if st == "selesai":
            p["selesai"] += 1
        else:
            p["overdue"] += 1

    current_by_pic = {}
    for loc in locations:
        if not loc["pic"]:
            continue
        c = current_by_pic.setdefault(
            loc["pic"], {"lokasiSaatIni": 0, "selesaiSaatIni": 0, "prosesSaatIni": 0, "overdueSaatIni": 0}
        )
        c["lokasiSaatIni"] += 1
        c[{"selesai": "selesaiSaatIni", "proses": "prosesSaatIni", "overdue": "overdueSaatIni"}[loc["status"]]] += 1

    leaderboard = []
    for p in pic_stats.values():
        pct = round(100 * p["selesai"] / p["lokasiDitangani"]) if p["lokasiDitangani"] else 0
        entry = {
            **p,
            "pctTepatWaktu": pct,
            "basisBulan": "2026-07",
            **current_by_pic.get(p["pic"], {"lokasiSaatIni": 0, "selesaiSaatIni": 0, "prosesSaatIni": 0, "overdueSaatIni": 0}),
        }
        leaderboard.append(entry)
    leaderboard.sort(key=lambda x: x["pctTepatWaktu"], reverse=True)

    pic_filled = sum(1 for loc in locations if loc["pic"])
    pic_coverage_pct = round(100 * pic_filled / len(locations)) if locations else 0

    tipe_counts = {"LAMA": 0, "BARU": 0, "TIDAK_DIKETAHUI": 0}
    for loc in locations:
        tipe_counts[loc["tipeClient"] or "TIDAK_DIKETAHUI"] += 1

    dataset = {
        "generatedAt": datetime.now().isoformat(timespec="seconds"),
        "sourceFile": src.split("/")[-1],
        "asOfDate": as_of,
        "currentMonth": current_code,
        "months": months_out,
        "totalLokasiAktif": len(locations),
        "totalLokasiDikecualikan": len(excluded),
        "tipeCounts": tipe_counts,
        "picCoveragePct": pic_coverage_pct,
        "locations": locations,
        "excludedLocations": excluded,
        "heatmap": heatmap_sites,
        "leaderboard": leaderboard,
    }

    out_path = "src/data/dashboard-data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(dataset, f, ensure_ascii=False, indent=2)
    print(f"wrote {out_path}")
    print(f"  lokasi aktif (bulan {current_code}): {len(locations)}")
    print(f"  dikecualikan: {len(excluded)}")
    print(f"  PIC coverage: {pic_coverage_pct}%")
    print(f"  leaderboard PIC: {len(leaderboard)} orang")
    print(f"  heatmap sites: {len(heatmap_sites)}")


if __name__ == "__main__":
    main()
