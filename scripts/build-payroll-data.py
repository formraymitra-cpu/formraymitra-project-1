#!/usr/bin/env python3
"""
Menggabungkan export "REKAP GAJI SEMUA LOKASI" dan "REKONSILIASI REKAP GAJI
DAN MUTASI" jadi satu dataset JSON untuk dashboard Rekonsiliasi Payroll
(src/data/payroll-data.json).

Jalankan ulang setiap ada export baru dari kedua spreadsheet:
    python3 scripts/build-payroll-data.py <rekap-gaji.xlsx> <rekonsiliasi.xlsx> [as-of-date YYYY-MM-DD]

Asumsi struktur (lihat README di apps-script-payroll/ untuk detail & cara
sinkron live lewat Apps Script):

REKAP GAJI SEMUA LOKASI - 1 sheet per bulan (nama sheet = nama bulan huruf
besar tanpa tahun, mis. "AGUSTUS"), header di baris 4, data mulai baris 5:
  NO | NAMA LOKASI | SHEET | PIC GAJI | NAMA REKAP GAJI | LINK GAJI PIC |
  KOLOM GAJI | KOLOM DITERIMA KARYAWAN | CEK | BANK | RAB | GAJI |
  DITERIMA KARYAWAN | BPJS KES | BPJS TK | PAYROLL | NOMINAL MUTASI |
  STATUS (K=L=M+N+O+P)

REKONSILIASI REKAP GAJI DAN MUTASI - beberapa sheet, yang dipakai di sini:
  PERIODE_GAJI, SUMBER_MUTASI, SUMBER_REKAP, HASIL_PENGECEKAN.
HASIL_PENGECEKAN adalah hasil fuzzy-matching per karyawan (REKAP vs MUTASI
bank) yang sudah dihitung oleh Apps Script yang sudah ada di spreadsheet
tsb — script ini HANYA membaca hasilnya, tidak menghitung ulang matching.
"""
import json
import re
import sys
from datetime import date, datetime

import openpyxl

TOLERANSI_RUPIAH = 5  # selisih di bawah ini dianggap pembulatan, bukan anomali

MONTH_ID = [
    "JANUARI", "FEBRUARI", "MARET", "APRIL", "MEI", "JUNI",
    "JULI", "AGUSTUS", "SEPTEMBER", "OKTOBER", "NOVEMBER", "DESEMBER",
]


def norm_lokasi(name):
    if not name:
        return ""
    s = str(name).upper().replace(".", "")
    s = re.sub(r"\bPENGECEKAN\b", "", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s


def num(v):
    if v is None or v == "":
        return None
    try:
        f = float(v)
    except (TypeError, ValueError):
        return None
    return f


def iso(v):
    if isinstance(v, datetime):
        return v.date().isoformat()
    if isinstance(v, date):
        return v.isoformat()
    return None


def clean(v):
    if v is None:
        return None
    s = str(v).strip()
    return s if s else None


def load(path):
    return openpyxl.load_workbook(path, data_only=True)


def parse_periode(wb):
    ws = wb["PERIODE_GAJI"]
    rows = [r for r in ws.iter_rows(min_row=2, values_only=True) if r[0]]
    aktif = next((r for r in rows if clean(r[4]) == "AKTIF"), rows[-1] if rows else None)
    if not aktif:
        return {"nama": None, "mulai": None, "selesai": None, "bulan": None, "tahun": None}
    nama = clean(aktif[1]) or ""
    parts = nama.split()
    bulan = next((p for p in parts if p in MONTH_ID), None)
    tahun = next((p for p in parts if p.isdigit() and len(p) == 4), None)
    return {
        "nama": nama,
        "mulai": iso(aktif[2]),
        "selesai": iso(aktif[3]),
        "bulan": bulan,
        "tahun": tahun,
    }


def parse_sumber_rekap_mutasi(wb):
    picImports = []
    if "SUMBER_MUTASI" in wb.sheetnames:
        ws = wb["SUMBER_MUTASI"]
        for r in ws.iter_rows(min_row=2, values_only=True):
            if not r[0]:
                continue
            picImports.append({
                "pic": clean(r[1]) or "",
                "namaFile": clean(r[2]) or "",
                "tanggalImport": iso(r[4]),
                "status": clean(r[6]) or "",
            })
    return picImports


def parse_rekap_lokasi(wb, sheet_name):
    if sheet_name not in wb.sheetnames:
        raise SystemExit(
            f"Sheet bulan '{sheet_name}' tidak ditemukan di REKAP GAJI SEMUA LOKASI. "
            f"Sheet yang ada: {wb.sheetnames}"
        )
    ws = wb[sheet_name]
    rows = []
    for r in ws.iter_rows(min_row=5, values_only=True):
        nama = clean(r[1]) if len(r) > 1 else None
        if not nama:
            continue
        rows.append({
            "nama": nama,
            "pic": clean(r[3]) if len(r) > 3 else None,
            "namaRekapGaji": clean(r[4]) if len(r) > 4 else None,
            "linkGajiPic": clean(r[5]) if len(r) > 5 else None,
            "bank": clean(r[9]) if len(r) > 9 else None,
            "rab": num(r[10]) if len(r) > 10 else None,
            "gaji": num(r[11]) if len(r) > 11 else None,
            "diterimaKaryawan": num(r[12]) if len(r) > 12 else None,
            "bpjsKes": num(r[13]) if len(r) > 13 else None,
            "bpjsTk": num(r[14]) if len(r) > 14 else None,
            "payroll": num(r[15]) if len(r) > 15 else None,
            "statusKomponen": clean(r[17]) if len(r) > 17 else None,
        })
    return rows


def monthly_totals(wb):
    out = []
    for sheet_name in wb.sheetnames:
        if sheet_name.upper() not in MONTH_ID:
            continue
        rows = parse_rekap_lokasi(wb, sheet_name)
        agg = {
            "code": sheet_name,
            "label": sheet_name.title(),
            "totalRab": sum(r["rab"] or 0 for r in rows),
            "totalGaji": sum(r["gaji"] or 0 for r in rows),
            "totalDiterimaKaryawan": sum(r["diterimaKaryawan"] or 0 for r in rows),
            "totalBpjsKes": sum(r["bpjsKes"] or 0 for r in rows),
            "totalBpjsTk": sum(r["bpjsTk"] or 0 for r in rows),
            "totalPayroll": sum(r["payroll"] or 0 for r in rows),
            "lokasiTerisi": sum(1 for r in rows if r["gaji"] is not None),
            "totalLokasi": len(rows),
        }
        out.append(agg)
    order = {m: i for i, m in enumerate(MONTH_ID)}
    out.sort(key=lambda a: order.get(a["code"].upper(), 99))
    return out


def parse_hasil_pengecekan(wb):
    if "HASIL_PENGECEKAN" not in wb.sheetnames:
        return {}
    ws = wb["HASIL_PENGECEKAN"]
    by_lokasi = {}
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r[0]:
            continue
        lokasi_raw = clean(r[2])
        if not lokasi_raw:
            continue
        acuan = clean(r[17])
        jumlah_ref = acuan.count("↔") + 1 if acuan else 0
        member = {
            "namaRekap": clean(r[3]) or "",
            "namaMutasi": clean(r[4]),
            "nominalRekap": num(r[5]),
            "nominalMutasi": num(r[6]),
            "selisih": num(r[7]),
            "bank": clean(r[8]),
            "tanggalMutasi": iso(r[9]),
            "statusAkhir": clean(r[16]),
            "transferGanda": jumlah_ref >= 2,
            "jumlahRefGanda": jumlah_ref,
            "konfirmasiManual": r[18] is True,
            "hasilManual": clean(r[19]),
        }
        by_lokasi.setdefault(lokasi_raw, []).append(member)
    return by_lokasi


def match_lokasi(rekap_nama, hasil_by_lokasi, hasil_norm_index, manual_map):
    if rekap_nama in manual_map:
        target = manual_map[rekap_nama]
        if target in hasil_by_lokasi:
            return target, "manual", []

    key = norm_lokasi(rekap_nama)
    exact = hasil_norm_index.get(key)
    if exact and len(exact) == 1:
        return exact[0], "auto", []

    candidates = set()
    for raw_lokasi, norm_key in [(raw, norm_lokasi(raw)) for raw in hasil_by_lokasi]:
        if not norm_key or not key:
            continue
        if norm_key in key or key in norm_key:
            candidates.add(raw_lokasi)
    if len(candidates) == 1:
        return next(iter(candidates)), "auto", []
    if len(candidates) > 1:
        return None, "ambigu", sorted(candidates)
    return None, "belum-ada", []


def parse_manual_mapping(wb):
    """Sheet opsional MAPPING_LOKASI di workbook REKAP GAJI: NAMA DI REKAP | NAMA DI REKONSILIASI."""
    out = {}
    if "MAPPING_LOKASI" not in wb.sheetnames:
        return out
    ws = wb["MAPPING_LOKASI"]
    for r in ws.iter_rows(min_row=2, values_only=True):
        if not r or not r[0] or not r[1]:
            continue
        out[clean(r[0])] = clean(r[1])
    return out


def build_locations(rekap_rows, hasil_by_lokasi, manual_map):
    hasil_norm_index = {}
    for raw in hasil_by_lokasi:
        hasil_norm_index.setdefault(norm_lokasi(raw), []).append(raw)

    raw_matches = []
    for row in rekap_rows:
        matched_raw, confidence, kandidat = match_lokasi(
            row["nama"], hasil_by_lokasi, hasil_norm_index, manual_map
        )
        raw_matches.append((row, matched_raw, confidence, kandidat))

    # Satu lokasi rekonsiliasi yang ke-klaim >1 baris REKAP (di luar mapping
    # manual yang eksplisit) berarti pencocokan otomatis tidak unik -> jangan
    # gabungkan datanya ke lebih dari satu lokasi (dobel-hitung), turunkan ke
    # "ambigu" supaya perlu diverifikasi manual (MAPPING_LOKASI).
    target_owners = {}
    for row, matched_raw, confidence, _ in raw_matches:
        if matched_raw and confidence != "manual":
            target_owners.setdefault(matched_raw, []).append(row["nama"])
    contested = {t for t, owners in target_owners.items() if len(owners) > 1}

    locations = []
    for row, matched_raw, confidence, kandidat in raw_matches:
        if matched_raw in contested and confidence != "manual":
            rekap_pesaing = sorted(n for n in set(target_owners[matched_raw]) if n != row["nama"])
            kandidat = [f"{matched_raw} (juga diklaim oleh: {', '.join(rekap_pesaing)})"]
            matched_raw, confidence = None, "ambigu"

        anggota = hasil_by_lokasi.get(matched_raw, []) if matched_raw else []

        jumlah_anggota = len(anggota)
        jumlah_sudah = sum(1 for a in anggota if a["nominalMutasi"] is not None)
        nominal_seharusnya = sum(a["nominalRekap"] or 0 for a in anggota)
        nominal_tertransfer = sum(a["nominalMutasi"] or 0 for a in anggota)

        komponen = (row["diterimaKaryawan"] or 0) + (row["bpjsKes"] or 0) + (row["bpjsTk"] or 0) + (row["payroll"] or 0)

        if jumlah_anggota == 0:
            status_role1 = "tanpa-data"
            selisih_role1 = None
        elif jumlah_sudah < jumlah_anggota:
            status_role1 = "proses"
            selisih_role1 = round(komponen - nominal_tertransfer, 2)
        else:
            selisih_role1 = round(komponen - nominal_tertransfer, 2)
            status_role1 = "sesuai" if abs(selisih_role1) <= TOLERANSI_RUPIAH else "selisih"

        if row["rab"] is None or row["gaji"] is None:
            status_role2 = "rab-kosong"
            selisih_role2 = None
        else:
            selisih_role2 = round(row["gaji"] - row["rab"], 2)
            status_role2 = "sesuai" if abs(selisih_role2) <= TOLERANSI_RUPIAH else "selisih"

        locations.append({
            "nama": row["nama"],
            "pic": row["pic"],
            "namaRekapGaji": row["namaRekapGaji"],
            "linkGajiPic": row["linkGajiPic"],
            "bank": row["bank"],
            "rab": row["rab"],
            "gaji": row["gaji"],
            "diterimaKaryawan": row["diterimaKaryawan"],
            "bpjsKes": row["bpjsKes"],
            "bpjsTk": row["bpjsTk"],
            "payroll": row["payroll"],
            "statusKomponen": row["statusKomponen"],
            "lokasiRekonsiliasi": matched_raw,
            "matchConfidence": confidence,
            "kandidatAmbigu": kandidat,
            "jumlahAnggota": jumlah_anggota,
            "jumlahSudahTertransfer": jumlah_sudah,
            "nominalSeharusnya": round(nominal_seharusnya, 2),
            "nominalTertransfer": round(nominal_tertransfer, 2),
            "statusRole1": status_role1,
            "selisihRole1": selisih_role1,
            "statusRole2": status_role2,
            "selisihRole2": selisih_role2,
            "anggota": anggota,
        })
    return locations


def build_notices(locations, hasil_by_lokasi):
    notices = []
    nid = 0

    def add(kategori, severity, lokasi, nama, nominal, ket):
        nonlocal nid
        nid += 1
        notices.append({
            "id": f"N{nid:04d}",
            "kategori": kategori,
            "severity": severity,
            "lokasi": lokasi,
            "namaKaryawan": nama,
            "nominalDampak": nominal,
            "keterangan": ket,
        })

    # Notice per-karyawan HARUS jalan dari hasil_by_lokasi (seluruh isi
    # HASIL_PENGECEKAN), bukan dari locations[].anggota -- karena tidak semua
    # lokasi di HASIL_PENGECEKAN punya padanan di REKAP GAJI SEMUA LOKASI
    # (mis. lokasi tipe SATPAM seperti "SAMSAT 2 PENGECEKAN" kadang tidak ada
    # baris REKAP-nya sama sekali). Kalau cuma baca locations[].anggota,
    # karyawan di lokasi begini (dan double-transfer/selisihnya) tidak akan
    # pernah muncul sebagai notice.
    label_by_raw = {
        l["lokasiRekonsiliasi"]: l["nama"]
        for l in locations
        if l["lokasiRekonsiliasi"] and l["matchConfidence"] in ("auto", "manual")
    }

    for raw_lokasi, anggota in hasil_by_lokasi.items():
        lokasi_label = label_by_raw.get(raw_lokasi, raw_lokasi)
        for m in anggota:
            if m["transferGanda"]:
                add(
                    "double-transfer", "tinggi", lokasi_label, m["namaRekap"],
                    m["nominalMutasi"],
                    f"{m['jumlahRefGanda']} baris mutasi bank cocok dengan nama ini — cek kemungkinan transfer ganda.",
                )
            if m["nominalMutasi"] is not None and m["selisih"] is not None and abs(m["selisih"]) > TOLERANSI_RUPIAH:
                if m["selisih"] > 0:
                    add(
                        "transfer-kurang", "tinggi", lokasi_label, m["namaRekap"], m["selisih"],
                        f"Nominal rekap Rp {m['nominalRekap']:,.0f} vs mutasi Rp {m['nominalMutasi']:,.0f} — transfer kurang Rp {m['selisih']:,.0f}.".replace(",", "."),
                    )
                else:
                    add(
                        "transfer-lebih", "tinggi", lokasi_label, m["namaRekap"], abs(m["selisih"]),
                        f"Nominal rekap Rp {m['nominalRekap']:,.0f} vs mutasi Rp {m['nominalMutasi']:,.0f} — transfer lebih Rp {abs(m['selisih']):,.0f}.".replace(",", "."),
                    )
            if m["statusAkhir"] and ("PERLU CEK" in m["statusAkhir"]):
                add(
                    "nama-lokasi-perlu-cek", "sedang", lokasi_label, m["namaRekap"], m["nominalRekap"],
                    f"Status pencocokan: {m['statusAkhir']}.",
                )
            if m["konfirmasiManual"] and m["nominalMutasi"] is None:
                # Admin sudah cek manual dan bilang "SESUAI", tapi kolom NOMINAL MUTASI di
                # HASIL_PENGECEKAN tetap kosong. Ini bisa berarti dua hal yang beda: (a) memang
                # belum tertransfer (dan admin sudah tahu itu), atau (b) sudah tertransfer tapi
                # gagal ke-link otomatis ke baris mutasinya (butuh dilengkapi manual di sheet
                # HASIL_PENGECEKAN supaya ikut kehitung sebagai "tertransfer"). Tidak bisa
                # dibedakan otomatis dari kolom yang ada -- makanya diberi notice terpisah,
                # bukan ditebak/digabung ke total tertransfer begitu saja.
                add(
                    "konfirmasi-tanpa-nominal", "sedang", lokasi_label, m["namaRekap"], m["nominalRekap"],
                    "Sudah dicek manual dan ditandai \"" + (m["hasilManual"] or "SESUAI") +
                    "\", tapi kolom NOMINAL MUTASI di HASIL_PENGECEKAN masih kosong sehingga tidak "
                    "ikut terhitung di Total Gaji Tertransfer. Cek lagi: kalau memang sudah "
                    "tertransfer, lengkapi nominal & tanggal mutasinya di HASIL_PENGECEKAN; kalau "
                    "memang belum, tidak perlu tindakan lebih lanjut.",
                )

    # Notice per-lokasi (butuh konteks RAB/GAJI/komponen dari REKAP) tetap
    # jalan dari `locations`, karena inherently REKAP-anchored.
    for loc in locations:
        if loc["jumlahAnggota"] > 0 and loc["jumlahSudahTertransfer"] < loc["jumlahAnggota"]:
            belum = loc["jumlahAnggota"] - loc["jumlahSudahTertransfer"]
            add(
                "belum-transfer", "sedang", loc["nama"], None, None,
                f"{belum} dari {loc['jumlahAnggota']} karyawan belum ditemukan padanan mutasinya (belum tertransfer / belum ter-import).",
            )

        if loc["matchConfidence"] == "belum-ada" and loc["diterimaKaryawan"]:
            add(
                "lokasi-tidak-ketemu", "sedang", loc["nama"], None, None,
                "Lokasi ini belum ditemukan di data rekonsiliasi (HASIL_PENGECEKAN) — rekonsiliasi mungkin belum dimulai untuk lokasi ini, atau nama lokasi berbeda antara kedua spreadsheet.",
            )

        if loc["statusRole2"] == "selisih":
            add(
                "gaji-vs-rab", "sedang", loc["nama"], None, abs(loc["selisihRole2"]),
                f"GAJI (Rp {loc['gaji']:,.0f}) tidak sama dengan RAB (Rp {loc['rab']:,.0f}) — cek apakah karena perubahan jumlah anggota (resign/belum ada pengganti) atau kesalahan input.".replace(",", "."),
            )

    severity_order = {"tinggi": 0, "sedang": 1}
    notices.sort(key=lambda n: (severity_order[n["severity"]], n["kategori"]))
    return notices


def build_kpi(locations, hasil_by_lokasi):
    total_gaji_tertransfer = sum(
        (m["nominalMutasi"] or 0) for anggota in hasil_by_lokasi.values() for m in anggota
    )
    total_rab = sum(l["rab"] or 0 for l in locations)
    total_gaji = sum(l["gaji"] or 0 for l in locations)
    total_diterima = sum(l["diterimaKaryawan"] or 0 for l in locations)
    total_bpjs_kes = sum(l["bpjsKes"] or 0 for l in locations)
    total_bpjs_tk = sum(l["bpjsTk"] or 0 for l in locations)
    total_payroll = sum(l["payroll"] or 0 for l in locations)
    lokasi_rab_terisi = sum(1 for l in locations if l["rab"] is not None)

    def komponen(l):
        return (l["diterimaKaryawan"] or 0) + (l["bpjsKes"] or 0) + (l["bpjsTk"] or 0) + (l["payroll"] or 0)

    terekonsiliasi = [l for l in locations if l["jumlahAnggota"] > 0]
    belum = [l for l in locations if l["jumlahAnggota"] == 0]
    kewajiban_terekonsiliasi = sum(komponen(l) for l in terekonsiliasi)
    kewajiban_belum = sum(komponen(l) for l in belum)

    return {
        "totalRab": round(total_rab, 2),
        "totalGaji": round(total_gaji, 2),
        "totalDiterimaKaryawan": round(total_diterima, 2),
        "totalBpjsKes": round(total_bpjs_kes, 2),
        "totalBpjsTk": round(total_bpjs_tk, 2),
        "totalPayroll": round(total_payroll, 2),
        "totalGajiTertransfer": round(total_gaji_tertransfer, 2),
        "kewajibanTerekonsiliasi": round(kewajiban_terekonsiliasi, 2),
        "kewajibanBelumRekonsiliasi": round(kewajiban_belum, 2),
        # selisih dihitung hanya atas lokasi yg SUDAH masuk proses rekonsiliasi (ada anggota
        # cocok di HASIL_PENGECEKAN) -- supaya gap "belum diproses" (wajar, periode masih
        # berjalan) tidak tercampur dengan gap "sudah diproses tapi nominalnya beda" (anomali).
        "selisihRole1": round(kewajiban_terekonsiliasi - total_gaji_tertransfer, 2),
        "selisihRole2": round(
            sum((l["gaji"] or 0) - (l["rab"] or 0) for l in locations if l["rab"] is not None), 2
        ),
        "lokasiRabTerisi": lokasi_rab_terisi,
    }


def main():
    if len(sys.argv) < 3:
        print(__doc__)
        sys.exit(1)
    rekap_path = sys.argv[1]
    rekon_path = sys.argv[2]
    as_of = sys.argv[3] if len(sys.argv) > 3 else date.today().isoformat()

    wb_rekap = load(rekap_path)
    wb_rekon = load(rekon_path)

    periode = parse_periode(wb_rekon)
    sheet_name = (periode["bulan"] or "").title().upper()
    if sheet_name not in wb_rekap.sheetnames:
        sheet_name = wb_rekap.sheetnames[-1]
        print(f"Peringatan: sheet bulan '{periode['bulan']}' tidak ketemu persis, pakai '{sheet_name}' (sheet terakhir).", file=sys.stderr)

    rekap_rows = parse_rekap_lokasi(wb_rekap, sheet_name)
    hasil_by_lokasi = parse_hasil_pengecekan(wb_rekon)
    manual_map = parse_manual_mapping(wb_rekap)
    pic_imports = parse_sumber_rekap_mutasi(wb_rekon)

    locations = build_locations(rekap_rows, hasil_by_lokasi, manual_map)
    notices = build_notices(locations, hasil_by_lokasi)
    kpi = build_kpi(locations, hasil_by_lokasi)
    totals = monthly_totals(wb_rekap)

    lokasi_sudah_dicek = sum(1 for l in locations if l["statusKomponen"] is not None)
    lokasi_sesuai = sum(1 for l in locations if l["statusKomponen"] == "SESUAI")
    lokasi_tidak_sesuai = sum(1 for l in locations if l["statusKomponen"] == "TIDAK SESUAI")
    lokasi_belum_rekon = sum(1 for l in locations if l["matchConfidence"] == "belum-ada")

    matched_auto = sum(1 for l in locations if l["matchConfidence"] == "auto")
    matched_ambigu = sum(1 for l in locations if l["matchConfidence"] == "ambigu")
    print(
        f"Lokasi: {len(locations)} | cocok otomatis ke rekonsiliasi: {matched_auto} | "
        f"ambigu (perlu MAPPING_LOKASI manual): {matched_ambigu} | belum ada data rekonsiliasi: {lokasi_belum_rekon}",
        file=sys.stderr,
    )

    dataset = {
        "generatedAt": datetime.now().isoformat(),
        "periode": periode["nama"],
        "periodeLabel": (periode["nama"] or "").title(),
        "periodeMulai": periode["mulai"],
        "periodeSelesai": periode["selesai"],
        "asOfDate": as_of,
        "kpi": kpi,
        "totalLokasi": len(locations),
        "lokasiSudahDicek": lokasi_sudah_dicek,
        "lokasiSesuai": lokasi_sesuai,
        "lokasiTidakSesuai": lokasi_tidak_sesuai,
        "lokasiBelumAdaRekonsiliasi": lokasi_belum_rekon,
        "picImports": pic_imports,
        "locations": locations,
        "notices": notices,
        "monthlyTotals": totals,
    }

    out_path = "src/data/payroll-data.json"
    with open(out_path, "w", encoding="utf-8") as f:
        json.dump(dataset, f, ensure_ascii=False, indent=2)
    print(f"Ditulis ke {out_path} ({len(locations)} lokasi, {len(notices)} notice).", file=sys.stderr)


if __name__ == "__main__":
    main()
